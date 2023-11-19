import openpyxl
from pathlib import Path


class XlsxReader:
    def __init__(self, path: Path):
        self.workbook = openpyxl.load_workbook(path)
        self.worksheet = self.workbook.active

        self.row_size = self.worksheet.max_row

    def read(self) -> list:
        result = []

        for row in range(1, self.row_size):
            data = {
                'question': str(self.worksheet.cell(row=row, column=1).value),
                'correct_answer': str(self.worksheet.cell(row=row, column=2).value),
                'answer1': str(self.worksheet.cell(row=row, column=3).value),
                'answer2': str(self.worksheet.cell(row=row, column=4).value),
                'answer3': str(self.worksheet.cell(row=row, column=5).value),
            }
            result.append(data)

        return result
